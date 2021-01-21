import requests
import re
from selenium import webdriver
import time
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
# 　　INTP
# 　　http://www.douban.com/group/20031/
# 　　INFP
# 　　http://www.douban.com/group/INFP/
# 　　INTJ
# 　　http://www.douban.com/group/intj/
# 　　INFJ
# 　　http://www.douban.com/group/119735/
# 　　ENTP
# 　　http://www.douban.com/group/116213/
# 　　ENFP
# 　　http://www.douban.com/group/ENFP/
# 　　ENTJ
# 　　http://www.douban.com/group/ENTJs/
# 　　ENFJ
# 　　http://www.douban.com/group/ENFJ/
# 　　
# 　　ISFJ
# 　　http://www.douban.com/group/108042/
# 　　ISTJ
# 　　http://www.douban.com/group/79036/
# 　　ISFP
# 　　http://www.douban.com/group/120688/
# 　　ISTP
# 　　http://www.douban.com/group/151685/
# 　　ESTP
# 　　http://www.douban.com/group/ESTP/
# 　　ESFP
# 　　http://www.douban.com/group/esfp/
# 　　ESFJ
# 　　http://www.douban.com/group/ESFJ/
# 　　ESTJ
# http://www.douban.com/group/ESTJ/　

information = []
wd = webdriver.Chrome(r'D:\chromedriver.exe')
wd.implicitly_wait(5)
all_address = []
url = 'http://www.douban.com/group/116213/'#您把上面12个网址一次复制到这里运行就可以了
name = []
headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36'}
def get_address(url):
    r = requests.get(url,headers=headers)
    txt = r.text
    adress = re.findall(r'a href="(.*?)" title=',txt)
    for i in adress:
        all_address.append(i)
    title = re.findall(r'''    <title>
        (.*?)
</title>''',txt)
    print(title)
    name.append(title[0])
    return all_address

address = get_address(url)
# print(all_address)
# print(len(all_address))


def get_message(url):
    wd.get(url)
    # time.sleep(3)

    #get url
    elements = wd.find_elements_by_css_selector('[class="topic-content"]')
    for element in elements:
        # 打印出元素对应的html
        # print(element.text)
        x = []
        x.append(element.text)
        information.append(element.text)
    replies = wd.find_elements_by_css_selector('[class=" reply-content"]')
    for reply in replies:
        # print(reply.text)
        y = []
        y.append(reply.text)
        information.append(reply.text)
    # r = requests.get(url,headers=headers)
    # txt = r.text
    # title = re.findall(r'title: "INTP",',txt)
    # print('giao',title)
# for i in all_address:
#     get_message(i)
for i in range(len(all_address)):
    print('一共%d个网页，正在爬取第%d个网页'%(len(all_address),i+1))
    get_message(all_address[i])

# for i in range(0,3):
#     get_message(all_address[i])

# print(information)
def main():
    print('saving...')
    wb = load_workbook(r'D:\model.xlsx')
    sheet = wb.active
    #获取行列数
    rows = sheet.max_row
    column = sheet.max_column
    for i in range(len(information)):
        try:
            sheet.cell(i+2,2).value = information[i]
            sheet.cell(i+2,1).value = name[0]
        except:
            continue
    # for i in range(len(information)):
    #     sheet1.cell(i+2, 2).value=information[i]
    wb.save(r'D:\{}.xlsx'.format(name[0]))
    data_xls = pd.read_excel(r'D:\{}.xlsx'.format(name[0]), index_col=0)
    data_xls.to_csv(r'D:\{}.csv'.format(name[0]), encoding='utf-8-sig')
    print('sucess!')

main()
wd.quit